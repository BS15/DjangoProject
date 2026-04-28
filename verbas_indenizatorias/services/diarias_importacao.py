"""Servicos canonicos de importacao em lote para diarias."""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from commons.shared.csv_import_utils import build_csv_dict_reader

from credores.models import Credor
from verbas_indenizatorias.constants import STATUS_VERBA_APROVADA, STATUS_VERBA_SOLICITADA
from verbas_indenizatorias.models import Diaria
from verbas_indenizatorias.services.documentos import gerar_e_anexar_pcd_diaria

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependencia opcional em tempo de desenvolvimento
    load_workbook = None


_COLUNAS_IMPORTACAO_DIARIAS = (
    "NOME_BENEFICIARIO",
    "DATA_SOLICITACAO",
    "DATA_SAIDA",
    "DATA_RETORNO",
    "CIDADE_ORIGEM",
    "CIDADE_DESTINO",
    "OBJETIVO",
    "TIPO_SOLICITACAO",
)


class DiariaCsvValidationError(Exception):
    """Erro de validacao de linha de diaria em importacao CSV."""


def _normalizar_coluna(coluna):
    return str(coluna or "").strip().upper()


def _valor_texto(row, coluna):
    valor = row.get(coluna, "")
    if valor is None:
        return ""
    return str(valor).strip()


def _parse_data(raw_value, line_num, nome_coluna):
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value

    valor = str(raw_value or "").strip()
    if not valor:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(valor, fmt).date()
        except ValueError:
            continue

    raise DiariaCsvValidationError(
        f"Linha {line_num}: {nome_coluna} invalida. Use o formato DD/MM/AAAA."
    )


def _calcular_quantidade_preview(data_saida, data_retorno, tipo_solicitacao):
    diferenca_dias = (data_retorno - data_saida).days
    if (tipo_solicitacao or "INICIAL").upper() == "INICIAL":
        return Decimal(diferenca_dias) + Decimal("0.5")
    return Decimal(diferenca_dias)


def _build_xlsx_rows(planilha_file):
    if load_workbook is None:
        raise RuntimeError("Dependencia openpyxl indisponivel para leitura de XLSX.")

    planilha_file.seek(0)
    workbook = load_workbook(filename=BytesIO(planilha_file.read()), read_only=True, data_only=True)
    aba = workbook.active
    linhas = aba.iter_rows(values_only=True)

    cabecalho_bruto = next(linhas, None)
    if not cabecalho_bruto:
        return None, "Cabecalho invalido. O arquivo XLSX esta vazio.", []

    colunas = [_normalizar_coluna(valor) for valor in cabecalho_bruto]
    faltantes = [col for col in _COLUNAS_IMPORTACAO_DIARIAS if col not in colunas]
    if faltantes:
        faltantes_str = ", ".join(faltantes)
        return None, f"Cabecalho invalido. Colunas ausentes: {faltantes_str}", []

    rows = []
    for indice, valores in enumerate(linhas, start=2):
        if not valores or all(valor in (None, "") for valor in valores):
            continue

        row = {colunas[pos]: valores[pos] if pos < len(valores) else "" for pos in range(len(colunas))}
        rows.append((indice, row))

    return rows, None, colunas


def _build_rows_from_file(planilha_file):
    nome_arquivo = (getattr(planilha_file, "name", "") or "").lower()

    if nome_arquivo.endswith(".xlsx"):
        rows, erro, _ = _build_xlsx_rows(planilha_file)
        return rows, erro

    reader, erro = build_csv_dict_reader(
        planilha_file,
        encodings=("utf-8-sig", "utf-8", "latin-1"),
        encoding_error_message="Nao foi possivel decodificar o CSV. Use UTF-8 ou Latin-1.",
        required_columns=_COLUNAS_IMPORTACAO_DIARIAS,
    )
    if erro:
        return None, erro

    return [(line_num, row) for line_num, row in enumerate(reader, start=2)], None


def _parse_diaria_row(row, line_num):
    nome = _valor_texto(row, "NOME_BENEFICIARIO")
    credor = Credor.objects.filter(nome__iexact=nome, tipo="PF").first() or Credor.objects.filter(
        nome__icontains=nome, tipo="PF"
    ).first()
    if not credor:
        raise DiariaCsvValidationError(f"Linha {line_num}: Beneficiario '{nome}' nao encontrado no sistema.")

    data_saida = _parse_data(row.get("DATA_SAIDA"), line_num, "DATA_SAIDA")
    data_retorno = _parse_data(row.get("DATA_RETORNO"), line_num, "DATA_RETORNO")
    if not data_saida or not data_retorno:
        raise DiariaCsvValidationError(
            f"Linha {line_num}: DATA_SAIDA e DATA_RETORNO sao obrigatorias no formato DD/MM/AAAA."
        )

    raw_solicitacao = row.get("DATA_SOLICITACAO")
    if raw_solicitacao:
        data_solicitacao = _parse_data(raw_solicitacao, line_num, "DATA_SOLICITACAO")
        if not data_solicitacao:
            data_solicitacao = datetime.today().date()
    else:
        data_solicitacao = datetime.today().date()

    if data_retorno < data_saida:
        raise DiariaCsvValidationError(
            f"Linha {line_num}: Data de retorno ({_valor_texto(row, 'DATA_RETORNO')}) nao pode ser anterior a data de saida ({_valor_texto(row, 'DATA_SAIDA')})."
        )

    tipo_solicitacao = (_valor_texto(row, "TIPO_SOLICITACAO") or "INICIAL").upper()
    qtd = _calcular_quantidade_preview(data_saida, data_retorno, tipo_solicitacao)

    return {
        "beneficiario_id": credor.id,
        "beneficiario_nome": credor.nome,
        "data_solicitacao": data_solicitacao.isoformat(),
        "data_saida": data_saida.isoformat(),
        "data_retorno": data_retorno.isoformat(),
        "quantidade_diarias": str(qtd),
        "cidade_origem": _valor_texto(row, "CIDADE_ORIGEM"),
        "cidade_destino": _valor_texto(row, "CIDADE_DESTINO"),
        "objetivo": _valor_texto(row, "OBJETIVO"),
        "tipo_solicitacao": tipo_solicitacao,
    }


def preview_diarias_lote(planilha_file):
    """Le arquivo XLSX/CSV e retorna preview serializavel com erros."""
    preview = []
    erros = []
    try:
        linhas, erro = _build_rows_from_file(planilha_file)
        if erro:
            return {"preview": [], "erros": [erro]}

        for line_num, row in linhas:
            try:
                preview.append(_parse_diaria_row(row, line_num))
            except DiariaCsvValidationError as exc:
                erros.append(str(exc))
    except Exception as exc:
        erros.append(f"Erro ao ler arquivo: {exc}")
    return {"preview": preview, "erros": erros}


def confirmar_diarias_lote(preview_items, usuario):
    """Compat: cria diarias em lote no modo padrao (solicitacao para autorizacao)."""
    return confirmar_diarias_lote_com_modo(preview_items, usuario, solicitacao_assinada=False)


def confirmar_diarias_lote_com_modo(preview_items, usuario, solicitacao_assinada=False):
    """Cria diarias em lote no modo padrao ou no modo de solicitacao ja assinada."""
    resultados = {"sucessos": 0, "erros": [], "modo_solicitacao_assinada": bool(solicitacao_assinada)}
    status_destino = STATUS_VERBA_APROVADA if solicitacao_assinada else STATUS_VERBA_SOLICITADA

    for item in preview_items:
        try:
            diaria = Diaria.objects.create(
                beneficiario_id=item["beneficiario_id"],
                proponente=usuario,
                data_solicitacao=date.fromisoformat(item["data_solicitacao"]),
                data_saida=date.fromisoformat(item["data_saida"]),
                data_retorno=date.fromisoformat(item["data_retorno"]),
                cidade_origem=item.get("cidade_origem", ""),
                cidade_destino=item.get("cidade_destino", ""),
                objetivo=item.get("objetivo", ""),
                tipo_solicitacao=item.get("tipo_solicitacao", "INICIAL"),
            )
            diaria.definir_status(status_destino, autorizada=bool(solicitacao_assinada))
            if solicitacao_assinada:
                gerar_e_anexar_pcd_diaria(diaria, criador=usuario)
            resultados["sucessos"] += 1
        except Exception as exc:
            resultados["erros"].append(f"{item.get('beneficiario_nome', '?')}: {exc}")

    return resultados


__all__ = [
    "DiariaCsvValidationError",
    "preview_diarias_lote",
    "confirmar_diarias_lote",
    "confirmar_diarias_lote_com_modo",
]
